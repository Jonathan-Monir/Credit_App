import pandas as pd
from openpyxl import Workbook
from datetime import timedelta
import streamlit as st
import openpyxl
from PIL import Image
import streamlit as st
import warnings
import numpy as np
# Add this line at the beginning of your script or function to ignore the warning
pd.options.mode.chained_assignment = None



# Add this line at the end of your script or function to re-enable the warning
pd.options.mode.chained_assignment = 'warn'

warnings.filterwarnings("ignore", category=UserWarning)
im = Image.open("bill.png")
st.set_page_config(
    page_title="Account Receivable Invoice PRO",
    page_icon=im,
    layout="wide",
)

st.markdown("""
    <h1 style='text-align: center; margin-bottom: -35px;'>
    Account Receivable Invoice PRO
    </h1>
    """, unsafe_allow_html=True
)

password_placeholder = st.empty()
password_error = st.empty()

import datetime
current_date = datetime.date.today()
formatted_date = current_date.strftime('%d-%m-%Y')

dd = ''.join(formatted_date.split('-')[:2])
ps = dd +"@0111Jo"

if "password" not in st.session_state:
    password = password_placeholder.text_input('Please enter a password', type='password')
    st.session_state["password"] = password
elif "password" in st.session_state and st.session_state["password"] != ps:
    password = password_placeholder.text_input('Please enter a password', type='password')
    
elif "password" in st.session_state and st.session_state["password"] == ps:
    password = ps
    
# "st.session_state object:", st.session_state
if password != ps and len(password) > 0:
    password_error.error('Incorrect password. Please try again.')
    
elif password == ps:
    st.session_state["password"] = password
    password_placeholder.empty()
    password_error.empty()

        
        
    

    if "uploaded file" not in st.session_state:
        uploaded_file = st.file_uploader("Choose a file")
        st.session_state["uploaded file"] = uploaded_file
        
    elif st.session_state["uploaded file"] is None:
        uploaded_file = st.file_uploader("Choose a file")
        st.session_state["uploaded file"] = uploaded_file

    else:
        uploaded_file = st.file_uploader("Choose a file")
        uploaded_file = st.session_state["uploaded file"]
        
    if uploaded_file is not None:
        
        workbook = openpyxl.load_workbook(uploaded_file)
        # Get the sheet names
        sheet_names = tuple([None] + workbook.sheetnames)
        st.session_state["sheet_names"] = sheet_names
        sheet_select = False
        if "statment" not in sheet_names:
            original_title = '<p style="font-size: 30px;">sheet selection</p>'
            st.markdown(original_title, unsafe_allow_html=True)
            sheet_select = True
            statment = st.selectbox(
            'Choose statment sheet',
            sheet_names)
        else:
            statment = "statment"

        if "contract" not in sheet_names:
            if sheet_select == False:
                original_title = '<p style="font-size: 30px;">sheet selection</p>'
                st.markdown(original_title, unsafe_allow_html=True)
            if "contract" in st.session_state:
                contract= st.selectbox(
                'Choose contract sheet',
                sheet_names, key="1",index = st.session_state["contract"])
                st.session_state["contract"] = sheet_names.index(contract)
            else:
                contract= st.selectbox(
                'Choose contract sheet',
                sheet_names, key="1")
                st.session_state["contract"] = sheet_names.index(contract)
        else:
            contract= "contract"
            st.session_state["contract"] = contract
         
        try:
            if "Unnamed: 0" in statment:
                statment.drop(columns = "Unnamed: 0", inplace = True)
        except TypeError as e:
            st.write("Please choose statment sheet first")
            # Handle the error gracefully, e.g., provide a default value or show an error message

        statment = pd.read_excel(uploaded_file,sheet_name=statment)
        con = pd.read_excel(uploaded_file,sheet_name= st.session_state["contract"])
        
        columns_select = False
        if "Amount-hotel" not in statment.columns:
            original_title = '<p style="font-size: 30px;">columns selection</p>'
            st.markdown(original_title, unsafe_allow_html=True)
            columns_select = True
            st.session_state["hotel"]= statment[st.selectbox(
            'Choose amount hotel column',
            statment.columns)]
            statment["hotel"] = st.session_state["hotel"]
            
        if "Rate code" not in statment.columns:
            if columns_select == False:
                original_title = '<p style="font-size: 30px;">columns selection</p>'
                st.markdown(original_title, unsafe_allow_html=True)
                
            st.session_state["Rate code"] = statment[st.selectbox(
            'Choose Room Type column',
            statment.columns)]
            statment["Rate code"] = st.session_state["Rate code"]
            
        if "Departure" not in statment.columns:
            if columns_select == False:
                original_title = '<p style="font-size: 30px;">columns selection</p>'
                st.markdown(original_title, unsafe_allow_html=True)
                
            st.session_state["Departure"] = statment[st.selectbox(
            'Choose Departure column',
            statment.columns)]
            statment["Departure"] = st.session_state["Departure"]
        if "Arrival" not in statment.columns:
            if not columns_select:
                original_title = '<p style="font-size: 30px;">columns selection</p>'
                st.markdown(original_title, unsafe_allow_html=True)
                
            statment["Arrival"] = statment[st.selectbox(
            'Choose Arrival column',
            statment.columns)]
            
        # preprocessing
        first_row = statment.iloc[0]
        num_nan = first_row.isnull().sum()
        shifted = False
        # Get the number of null values in each row
        statment = statment [~statment['Arrival'].isnull()]
        
        # Drop rows with more than 3 null values
        if statment['Arrival'].dtype == 'object':
            statment['Arrival'] = pd.to_datetime(statment['Arrival'])
        if statment['Departure'].dtype == 'object':
            statment['Departure'] = pd.to_datetime(statment['Departure'])
        # Assuming you have already loaded the dataset into a DataFrame named 'statment'

        # Check if any column contains null (NaN) in all its values
        columns_with_all_nulls = statment.isnull().all()

        # Get the column names to drop (where all values are NaN)
        columns_to_drop = columns_with_all_nulls[columns_with_all_nulls].index

        # Drop the columns with all nulls from the DataFrame
        statment.drop(columns=columns_to_drop, inplace=True)

        # Print the DataFrame after dropping the columns

        while num_nan >= 6:
            shifted = True
            statment = statment.drop(index=0).reset_index(drop=True)
            first_row = statment.iloc[0]
            num_nan = first_row.isnull().sum()
            
        if shifted == True:
            # Set the first row as the column names
            statment.columns = statment.iloc[0]

            # Drop the first row (old column names) from the DataFrame
            statment.drop(index=0, inplace=True)

            # Reset the index after dropping the first row
            statment.reset_index(drop=True, inplace=True)

        # sheet names
        statment.fillna(0)
        con.fillna(0)
        
        columns_to_check = ['Booking No.', 'Guest Name', 'Invoice', 'Arrival','Description']
        any_column_exists = any(col in statment.columns for col in columns_to_check)



        statment["other_price"]=0
        statment["UnNeeded_price"]=0
        statment["Total price currency"]=0
        
        SPO_name = st.selectbox(
            'Choose special offer sheet',
            sheet_names, key="0")
        if SPO_name == None:
            checked1 = st.checkbox("SPO by arrival",value = False)
            if checked1:
                if "spo by arrival" not in sheet_names:
                    original_title = '<p style="font-size: 30px;">spo by arrival selection</p>'
                    st.markdown(original_title, unsafe_allow_html=True)
                    sheet_select = True
                    spo_arrival = st.selectbox(
                    'Choose spo by arrival sheet sheet',
                    sheet_names)
                else:
                    spo_arrival = "spo by arrival"
                spo_arrival_df = pd.read_excel(uploaded_file,sheet_name=spo_arrival)
        
            for i in range(len(statment["Arrival"])): # loop
                
                
                
                Summing=0
                date_arrival = pd.Timestamp(statment.loc[i, "Arrival"])
                date_departure = pd.Timestamp(statment.loc[i, "Departure"])

                rate_code = statment["Rate code"][i]
                arrival_row = con[(con["first date"]<=date_arrival) & (con["second date"]>=date_arrival)]
                
                departure_row = con[(con["first date"]<=date_departure) & (con["second date"]>=date_departure)]
                # if arrival_row.isnull().any().item():
                #     break
                
                price_arrival_night = arrival_row[rate_code]
                
                date1_arrival = pd.to_datetime(arrival_row["first date"].values[0])
                date2_arrival = pd.to_datetime(arrival_row["second date"].values[0])
                
                date1_departure = pd.to_datetime(departure_row["first date"].values[0])
                date2_departure = pd.to_datetime(departure_row["second date"].values[0])

                if date_departure <= date2_arrival:
                    price = float(price_arrival_night * ((date_departure-date_arrival).days))
                    statment.loc[i, "Total price currency"] = price
                    
                else:
                    date_range = con[(date_arrival<=con["second date"]) & (date_departure>=con["first date"])]
                    diff = (date_range["second date"] - date_range["first date"]).dt.days 
                    diff += 1
                    for j in range(len(date_range[rate_code])):
                        Summing = (date_range[rate_code].iloc[j]*diff.iloc[j]) + Summing
                        
                    other_price = (((date_arrival-date1_arrival).days) * (arrival_row[rate_code].values[0]) + ((date2_departure-date_departure).days+1) * (departure_row[rate_code].values[0]))
                    statment.loc[i,"other_price"] = other_price
                    statment.loc[i,"Total price currency"] = Summing - other_price
                if checked1:
                    res_date = statment["Res_date"][i]
                    if ((res_date >= spo_arrival_df["first date"][0]) and (res_date <= spo_arrival_df["second date"][0])) and ((date_arrival >= spo_arrival_df["first date"][0]) and (date_arrival <= spo_arrival_df["second date"][0])):
                        night_price = spo_arrival_df[rate_code][0]
                        
                        nights = statment["Departure"][i] - statment["Arrival"][i]
                        statment.loc[i, "Total price currency"] = float(night_price * nights.days)
                    elif ((res_date >= spo_arrival_df["first date"][0]) and (res_date <= spo_arrival_df["second date"][0])) and ((date_arrival >= spo_arrival_df["first date"][1]) and (date_arrival <= spo_arrival_df["second date"][1])):
                        night_price = spo_arrival_df[rate_code][1]
                        nights = statment["Departure"][i] - statment["Arrival"][i]
                        statment.loc[i, "Total price currency"] = float(night_price * nights.days)
                
        else:
            
            SPO2_name = st.selectbox(
            'Choose another special offer.',
            sheet_names)
            spo = pd.read_excel(uploaded_file,sheet_name=SPO_name)
            if "other_price" not in statment:
                statment["other_price"]=0
            if "UnNeeded_price" not in statment:
                statment["UnNeeded_price"]=0
            if "Total price currency" not in statment:
                statment["Total price currency"]=0
            
            first_day_spo = pd.Timestamp(spo["first date"][0])
            last_day_spo = pd.Timestamp(spo["second date"].iloc[-1])
            hide_checkbox_label_style = """
                <style>
                    .checkbox-container .stCheckbox>label {
                        display: none;
                    }
                </style>
            """
            
            # Display the CSS
            st.markdown(hide_checkbox_label_style, unsafe_allow_html=True)
            # Create the checkbox
            checked1 = st.checkbox("SPO by arrival",value = False)
            if checked1:
                if "spo by arrival" not in sheet_names:
                    original_title = '<p style="font-size: 30px;">spo by arrival selection</p>'
                    st.markdown(original_title, unsafe_allow_html=True)
                    sheet_select = True
                    spo_arrival = st.selectbox(
                    'Choose spo by arrival sheet sheet',
                    sheet_names)
                else:
                    spo_arrival = "spo by arrival"
                spo_arrival_df = pd.read_excel(uploaded_file,sheet_name=spo_arrival)
                
            
            for i in range(len(statment["Arrival"])): # loop
                Summing=0
                date_arrival = pd.Timestamp(statment.loc[i, "Arrival"])
                date_departure = pd.Timestamp(statment.loc[i, "Departure"] - timedelta(1))

                res_date = statment["Res_date"][i]
                rate_code = statment["Rate code"][i]


                if  res_date < first_day_spo or res_date > last_day_spo:
                    
                    arrival_row = con[(con["first date"]<=date_arrival) & (con["second date"]>=date_arrival)]
                    departure_row = con[(con["first date"]<=date_departure) & (con["second date"]>=date_departure)]
                    
                    price_arrival_night = arrival_row[rate_code]
                    date1_arrival = pd.to_datetime(arrival_row["first date"].values[0])
                    
                    date2_arrival = pd.to_datetime(arrival_row["second date"].values[0])
                    date1_departure = pd.to_datetime(departure_row["first date"].values[0])
                    date2_departure = pd.to_datetime(departure_row["second date"].values[0])
                    
                    if date_departure <= date2_arrival:
                        price = float(price_arrival_night * ((date_departure-date_arrival).days +1))
                        statment.loc[i, "Total price currency"] = price
                        
                    else:
                        date_range = con[(date_arrival<=con["first date"]) & (date_departure>=con["second date"])]
                        
                        diff = (date_range["second date"] - date_range["first date"]).dt.days + 1

                        for j in range(len(date_range[rate_code])):
                            Summing = (date_range[rate_code].iloc[j]*diff.iloc[j]) + Summing
                            
                            
                        other_price = (((date2_arrival-date_arrival).days+1) * (arrival_row[rate_code].values[0]) + ((date_departure-date1_departure).days+1) * (departure_row[rate_code].values[0]))
                        
                        statment.loc[i, "Total price currency"] = Summing + other_price
                else:
                    arrival_row = spo[(spo["first date"]<=date_arrival) & (spo["second date"]>=date_arrival)]
                    departure_row = spo[(spo["first date"]<=date_departure) & (spo["second date"]>=date_departure)]
                    
                    price_arrival_night = arrival_row[rate_code]
                    
                    
                    date1_arrival = pd.to_datetime(arrival_row["first date"].values[0])
                    date2_arrival = pd.to_datetime(arrival_row["second date"].values[0])
                    if date_departure<=last_day_spo:
                        
                        date1_departure = pd.to_datetime(departure_row["first date"].values[0])
                        date2_departure = pd.to_datetime(departure_row["second date"].values[0])
                        
                        if date_departure <= date2_arrival:
                            price = price_arrival_night * ((date_departure-date_arrival).days +1)
                            statment.loc[i, "Total price currency"] = price
                        else:
                            date_range = spo[(date_arrival<=spo["first date"]) & (date_departure>=spo["second date"])]
                            diff = (date_range["second date"] - date_range["first date"]).dt.days + 1

                            for j in range(len(date_range[rate_code])):
                                Summing = (date_range[rate_code].iloc[j]*diff.iloc[j]) + Summing
                                
                            other_price = (((date2_arrival-date_arrival).days+1) * (arrival_row[rate_code].values[0]) + ((date_departure-date1_departure).days+1) * (departure_row[rate_code].values[0]))
                            
                            
                            statment.loc[i, "Total price currency"] = Summing + other_price        
                    else:
                        departure_row = spo[(spo["first date"]<=last_day_spo) & (spo["second date"]>=last_day_spo)]

                        date1_arrival = pd.to_datetime(arrival_row["first date"].values[0])
                        date2_arrival = pd.to_datetime(arrival_row["second date"].values[0])  
                        
                        date1_departure = pd.to_datetime(departure_row["first date"].values[0])
                        date2_departure = pd.to_datetime(departure_row["second date"].values[0])
                        
                        if last_day_spo <= date2_arrival:
                            spo_part_price = price_arrival_night * ((last_day_spo-date_arrival).days +1)
                            spo_part_price=spo_part_price.values[0]
                        else:
                            date_range = spo[(date_arrival<=spo["first date"]) & (last_day_spo>=spo["second date"])]
                            diff = (date_range["second date"] - date_range["first date"]).dt.days + 1

                            for j in range(len(date_range[rate_code])):
                                Summing = (date_range[rate_code].iloc[j]*diff.iloc[j]) + Summing
                                
                            other_price = (((date2_arrival-date_arrival).days+1) * (arrival_row[rate_code].values[0]) + ((last_day_spo-date1_departure).days+1) * (departure_row[rate_code].values[0]))
                            
                            
                            spo_part_price = Summing + other_price  

                        date_arrival = last_day_spo + timedelta(1)

                        arrival_row = con[(con["first date"]<=date_arrival) & (con["second date"]>=date_arrival)]
                        departure_row = con[(con["first date"]<=date_departure) & (con["second date"]>=date_departure)]

                        price_arrival_night = arrival_row[rate_code]
                        date1_arrival = pd.to_datetime(arrival_row["first date"].values[0])
                        date2_arrival = pd.to_datetime(arrival_row["second date"].values[0])
                        
                        date1_departure = pd.to_datetime(departure_row["first date"].values[0])
                        date2_departure = pd.to_datetime(departure_row["second date"].values[0])
                        
                        if date_departure <= date2_arrival:
                            con_part_price = price_arrival_night * ((date_departure-date_arrival).days +1)
                            con_part_price=con_part_price.values[0]
                        else:
                            date_range = con[(date_arrival<=con["first date"]) & (date_departure>=con["second date"])]
                            diff = (date_range["second date"] - date_range["first date"]).dt.days + 1

                            for j in range(len(date_range[rate_code])):
                                Summing = (date_range[rate_code].iloc[j]*diff.iloc[j]) + Summing
                                
                            other_price = (((date2_arrival-date_arrival).days+1) * (arrival_row[rate_code].values[0]) + ((date_departure-date1_departure).days+1) * (departure_row[rate_code].values[0]))
                            
                            con_part_price = Summing + other_price
                            
                        statment.loc[i, "Total price currency"]=con_part_price + spo_part_price
                if checked1:        
                    if res_date >= spo_arrival_df["first date"][0] and res_date <= spo_arrival_df["second date"][0]:
                        if date_arrival >= spo_arrival_df["first date"][0] and date_arrival <= spo_arrival_df["second date"][0]:
                            price_arrival_night = spo_arrival_df[rate_code][0]
                            price = price_arrival_night * ((date_departure-date_arrival).days +1)
                            statment.loc[i, "Total price currency"] = price    
                        if date_arrival >= spo_arrival_df["first date"][1] and date_arrival <= spo_arrival_df["second date"][1]:
                            price_arrival_night = spo_arrival_df[rate_code][1]
                            price = price_arrival_night * ((date_departure-date_arrival).days +1)
                            statment.loc[i, "Total price currency"] = price
                    

                if SPO2_name:
                    Summing = 0
                    spo2 = pd.read_excel(uploaded_file,sheet_name=SPO2_name)
                    first_day_spo2 = spo2["first date"][0]
                    last_day_spo2 = spo2["second date"].iloc[-1]
                    if  res_date >= first_day_spo2 and res_date <= last_day_spo2:
                        date_arrival = pd.Timestamp(statment.loc[i, "Arrival"])
                        date_departure = pd.Timestamp(statment.loc[i, "Departure"]) - timedelta(1)

                        res_date = statment["Res_date"][i]
                        rate_code = statment["Rate code"][i]
                        arrival_row = spo2[(spo2["first date"]<=date_arrival) & (spo2["second date"]>=date_arrival)]
                        departure_row = spo2[(spo2["first date"]<=date_departure) & (spo2["second date"]>=date_departure)]
                        
                        price_arrival_night = arrival_row[rate_code]
                        date1_arrival = pd.to_datetime(arrival_row["first date"].values[0])
                        date2_arrival = pd.to_datetime(arrival_row["second date"].values[0])
                        
                        if date_departure<=last_day_spo2:

                            date1_departure = pd.to_datetime(departure_row["first date"].values[0])
                            date2_departure = pd.to_datetime(departure_row["second date"].values[0])
                            
                            if date_departure <= date2_arrival:
                                price = price_arrival_night * ((date_departure-date_arrival).days +1)
                                statment.loc[i, "Total price currency"] = price
                            else:
                                date_range = spo2[(date_arrival<=spo2["second date"]) & (date_departure>=spo2["first date"])]
                                diff = (date_range["second date"] - date_range["first date"]).dt.days + 1

                                for j in range(len(date_range[rate_code])):
                                    Summing = (date_range[rate_code].iloc[j]*diff.iloc[j]) + Summing
                                other_price = (((date_arrival-date1_arrival).days) * (arrival_row[rate_code].values[0]) + ((date2_departure-date_departure).days) * (departure_row[rate_code].values[0]))
                                
                                
                                statment.loc[i, "Total price currency"] = Summing - other_price      
                                # CONTINUE TO CONTRACT OR SPO HEREEEEE  
                        else:
                            departure_row = spo2[(spo2["first date"]<=last_day_spo) & (spo2["second date"]>=last_day_spo)]

                            date1_arrival = pd.to_datetime(arrival_row["first date"].values[0])
                            date2_arrival = pd.to_datetime(arrival_row["second date"].values[0])  
                            
                            date1_departure = pd.to_datetime(departure_row["first date"].values[0])
                            date2_departure = pd.to_datetime(departure_row["second date"].values[0])

                            if last_day_spo <= date2_arrival:
                                spo_part_price = price_arrival_night * ((last_day_spo-date_arrival).days +1)
                                spo_part_price=spo_part_price.values[0]
                            else:
                                date_range = spo2[(date_arrival<=spo2["first date"]) & (last_day_spo>=spo2["second date"])]
                                diff = (date_range["second date"] - date_range["first date"]).dt.days + 1

                                for j in range(len(date_range[rate_code])):
                                    Summing = (date_range[rate_code].iloc[j]*diff.iloc[j]) + Summing
                                    
                                other_price = (((date2_arrival-date_arrival).days+1) * (arrival_row[rate_code].values[0]) + ((last_day_spo-date1_departure).days+1) * (departure_row[rate_code].values[0]))
                                
                                
                                spo_part_price = Summing + other_price  

                            date_arrival = last_day_spo + timedelta(1)

                            arrival_row = con[(con["first date"]<=date_arrival) & (con["second date"]>=date_arrival)]
                            departure_row = con[(con["first date"]<=date_departure) & (con["second date"]>=date_departure)]

                            price_arrival_night = arrival_row[rate_code]

                            date1_arrival = pd.to_datetime(arrival_row["first date"].values[0])
                            date2_arrival = pd.to_datetime(arrival_row["second date"].values[0])
                            
                            date1_departure = pd.to_datetime(departure_row["first date"].values[0])
                            date2_departure = pd.to_datetime(departure_row["second date"].values[0])

                            if date_departure <= date2_arrival:
                                con_part_price = price_arrival_night * ((date_departure-date_arrival).days +1)
                                con_part_price=con_part_price.values[0]
                            else:
                                date_range = con[(date_arrival<=con["first date"]) & (date_departure>=con["second date"])]
                                diff = (date_range["second date"] - date_range["first date"]).dt.days + 1

                                for j in range(len(date_range[rate_code])):
                                    Summing = (date_range[rate_code].iloc[j]*diff.iloc[j]) + Summing
                                    
                                other_price = (((date2_arrival-date_arrival).days+1) * (arrival_row[rate_code].values[0]) + ((date_departure-date1_departure).days+1) * (departure_row[rate_code].values[0]))
                                
                                con_part_price = Summing + other_price
                                
                            statment.loc[i, "Total price currency"] =con_part_price + spo_part_price
        
          
        # Functions
        def F_eb1():
            statment.loc[i, "Total price currency"] -= (statment.loc[i, "Total price currency"] * (Offers_dict["eb1 percentage"]/100))  
            
        def F_eb2():
            statment.loc[i, "Total price currency"] -= (statment.loc[i, "Total price currency"] * (Offers_dict["eb2 percentage"]/100))
            
        def LT():
            statment.loc[i, "Total price currency"] -= ((statment.loc[i, "Total price currency"] * (Offers_dict["lt percentage"]/100)))
        
        def reduc():
            statment.loc[i, "Total price currency"] -= ((statment.loc[i, "Total price currency"] * (Offers_dict["reduc percentage"]/100)))
            
        def reduc2():
            statment.loc[i, "Total price currency"] -= ((statment.loc[i, "Total price currency"] * (Offers_dict["reduc2 percentage"]/100)))
            
        def offer(i,per):
            statment.loc[i, "Total price currency"] -= (statment.loc[i, "Total price currency"] * (per/100))  
            
        def offer_con(price,per):
            price -= (price * (per/100))  
            return price
        
        senior_bool = False
        if "Offers_dict" in st.session_state:
            Offers_dict = st.session_state["Offers_dict"]
            if "senior" in Offers_dict:
                senior_bool = True
                
        if "Spo_dict" in st.session_state:
            Spo_dict = st.session_state["Spo_dict"]
            if any(Spo_dict['senior']):
                senior_bool = True  
                  
        if senior_bool:
            sen_column = st.selectbox('please select senior column',options=statment.columns.insert(0,None))
            
        # Here are the other new offers
        if "Offers_dict" in st.session_state:
            Offers_dict = st.session_state["Offers_dict"]
            s_error = 0
                    
                            
            for i in range(len(statment["Arrival"])):
                
                # senior
                if "senior" in Offers_dict:
                    if Offers_dict["senior"] and (sen_column in statment.columns):
                        if (statment[sen_column][i] >0) :
                            Type_of_room = statment["Rate code"][i][0].lower()
                            
                            # adjusting room type
                            
                            type_of_room_mapping = {
                                "s": 1,
                                "d": 2,
                                "t": 3,
                                "q": 4
                            }

                            if Type_of_room in type_of_room_mapping:
                                mapped_value = type_of_room_mapping[Type_of_room]

                            
                            Total_price = statment.loc[i, "Total price currency"] * (statment[sen_column][i]/int(mapped_value)) * -(Offers_dict["senior percentage"]/100)
                            statment.loc[i, "Total price currency"] += Total_price
                            
                            
                # Early booking 1
                if "eb1" in Offers_dict:
                    if Offers_dict["eb1"]:
                        if statment["Res_date"][i] <= pd.Timestamp(Offers_dict["eb1 date"]):
                            F_eb1()
                            
                # Early booking 2
                if "eb2" in Offers_dict:
                    if Offers_dict["eb2"]:
                        if (statment["Res_date"][i] <= pd.Timestamp(Offers_dict["eb2 date"]) and statment["Res_date"][i] > pd.Timestamp(Offers_dict["eb1 date"])):
                            F_eb2()
                            
                if "FormSubmitter:Combinations-Submit" in st.session_state:
                    if st.session_state["FormSubmitter:Combinations-Submit"]:
                        Combin_dict = st.session_state['Combin_dict']
                        
                # long term
                if "lt" in Offers_dict:
                    if (Offers_dict["lt"]):
                        if "FormSubmitter:Combinations-Submit" not in st.session_state:
                            if ((statment["Departure"][i] - statment["Arrival"][i]).days > Offers_dict["lt days"]):
                                LT()
                        else:
                            if Combin_dict["long term_combin"]:
                                if ((statment["Departure"][i] - statment["Arrival"][i]).days > Offers_dict["lt days"]):
                                    LT()
                                    
                            
                # reduction 1
                if "reduc" in Offers_dict:
                    if (Offers_dict["reduc"]):
                        if "FormSubmitter:Combinations-Submit" not in st.session_state:
                            reduc()
                    else:
                        if "reduction_combin" in Offers_dict:
                            if Offers_dict["reduction_combin"]:
                                reduc()
                        
                # reduction 2
                if "reduc2" in Offers_dict:
                    if (Offers_dict["reduc2"]):
                        if "FormSubmitter:Combinations-Submit" not in Offers_dict:
                            reduc2()
                        else:
                            if "reduction2_combin" in Combin_dict:
                                if Combin_dict["reduction2_combin"]:
                                    reduc2()
                                
        # SPO's ******************************
        senior_cl = True
        def ceb1(cell,Spo_dict):
            if Spo_dict['eb1'][spo_num]:
                return (cell['Res_date'] < pd.Timestamp(Spo_dict['eb1 date'][spo_num])) and (Spo_dict['eb1'][spo_num])
            else:
                return False
            
        def ceb2(cell,Spo_dict):
            if Spo_dict['eb2'][spo_num]:
                return not(ceb1(cell, Spo_dict)) and(cell['Res_date'] < pd.Timestamp(Spo_dict['eb2 date'][spo_num]))
            else:
                return False
            
        def csenior(cell,Spo_dict):
            global senior_cl
            if Spo_dict['senior'][spo_num]:
                try:
                    return (Spo_dict['senior'][spo_num]) and (cell[sen_column] > 0)
                except:
                    if senior_cl:
                        st.error('Please add senior column first')
                        senior_cl = False
                    return False
            
                    
            else:
                return False
            
        def clt(cell,Spo_dict):
            if Spo_dict['long term'][spo_num]:
                return (Spo_dict['long term'][spo_num]) and ((cell['Departure']-cell['Arrival']).days>Spo_dict['lt days'][spo_num])
            else:
                return False
            
        def creduc1(cell,Spo_dict):
            if Spo_dict['reduc1'][spo_num]:
                return (Spo_dict['reduc1'][spo_num])
            else:
                return False
            
        def creduc2(cell,Spo_dict):
            if Spo_dict['reduc2'][spo_num]:
                return (creduc1) and (Spo_dict['reduc2'][spo_num])
            else:
                return False

        def calculate_offer(cell,Spo_dict,price,spo_num):
            if ceb1(cell,Spo_dict):
                price = price * (1 - (Spo_dict['eb1 percentage'][spo_num]/100))
                
            if ceb2(cell,Spo_dict):
                price = price * (1 - (Spo_dict['eb1 percentage'][spo_num]/100))
            try:
                csenior(cell,Spo_dict)
            except:
                st.error('Please Choose senior column first.')
            if csenior(cell,Spo_dict):
                type_of_room_mapping = {
                                "s": 1,
                                "d": 2,
                                "t": 3,
                                "q": 4
                            }
                if Type_of_room in type_of_room_mapping:
                                mapped_value = type_of_room_mapping[Type_of_room]
                
                price = price * (1-(statment[sen_column][guest]/int(mapped_value)) * (Spo_dict["senior percentage"][spo_num]/100))
                
            if clt(cell,Spo_dict):
                price = price * (1 - (Spo_dict['lt percentage'][spo_num]/100))

            if creduc1(cell,Spo_dict):
                price = price * (1 - (Spo_dict['reduc1 percentage'][spo_num]/100))
                
            if creduc2(cell,Spo_dict):
                price = price * (1 - (Spo_dict['reduc2 percentage'][spo_num]/100))
            return price
        
        def calculate_offer_con(cell,price):
                        
                        
            # Early booking 1
            if "eb1" in Offers_dict:
                if Offers_dict["eb1"]:
                    if cell["Res_date"] <= Offers_dict["eb1 date"]:
                        price = offer_con(price,Offers_dict['eb1 percentage'])
                        
            # Early booking 2
            if "eb2" in Offers_dict:
                if Offers_dict["eb2"]:
                    if (cell["Res_date"] <= Offers_dict["eb2 date"] and cell["Res_date"] > Offers_dict["eb1 date"]):
                        price = offer_con(price,Offers_dict['eb2 percentage'])
                        
            if "FormSubmitter:Combinations-Submit" in st.session_state:
                if st.session_state["FormSubmitter:Combinations-Submit"]:
                    Combin_dict = st.session_state['Combin_dict']
                    
            #senior
            if "senior" in Offers_dict:
                if Offers_dict["senior"]:
                    if cell[sen_column] >0:
                        Type_of_room = cell["Rate code"][0].lower()
                        
                        # adjusting room type
                        
                        type_of_room_mapping = {
                            "s": 1,
                            "d": 2,
                            "t": 3,
                            "q": 4
                        }

                        if Type_of_room in type_of_room_mapping:
                            mapped_value = type_of_room_mapping[Type_of_room]

                        
                        Total_price = price * (cell[sen_column]/int(mapped_value)) * -(Offers_dict["senior percentage"]/100)
                        price += Total_price
                        
                        
            # long term
            if "lt" in Offers_dict:
                if (Offers_dict["lt"]):
                    if "FormSubmitter:Combinations-Submit" not in st.session_state:
                        if ((cell["Departure"] - cell["Arrival"]).days > Offers_dict["lt days"]):
                            price = offer_con(price,Offers_dict['lt percentage'])
                    else:
                        if Combin_dict["long term_combin"]:
                            if ((cell["Departure"] - cell["Arrival"]).days > Offers_dict["lt days"]):
                                price = offer_con(price,Offers_dict['lt percentage'])
                                
                        
            # reduction 1
            if "reduc" in Offers_dict:
                if (Offers_dict["reduc"]):
                    if "FormSubmitter:Combinations-Submit" not in st.session_state:
                        price = offer_con(price,Offers_dict['reduc1 percentage'])
                else:
                    if "reduction_combin" in Offers_dict:
                        if Offers_dict["reduction_combin"]:
                            price = offer_con(price,Offers_dict['reduc1 percentage'])

                    
            # reduction 2
            if "reduc2" in Offers_dict:
                if (Offers_dict["reduc2"]):
                    if "FormSubmitter:Combinations-Submit" not in Offers_dict:
                        price = offer_con(price,Offers_dict['reduc2 percentage'])

                    else:
                        if "reduction2_combin" in Combin_dict:
                            if Combin_dict["reduction2_combin"]:
                                price = offer_con(price,Offers_dict['reduc2 percentage'])

            return price
        
        def append_and_pad(original_array, new_array):
            if original_array.ndim == 1:
                original_array = [original_array]
            # Find the maximum length between the two arrays
            max_length = max(len(original_array[0]), len(new_array))
            
            # Pad both arrays with zeros to match the maximum length
            original_array_padded = np.pad(original_array, ((0, 0), (0, max_length - len(original_array[0]))), 'constant')
            new_array_padded = np.pad(new_array, (0, max_length - len(new_array)), 'constant')
            
            # Stack the two arrays vertically
            result = np.vstack((original_array_padded, new_array_padded))
            
            return result
        def contract_removal(con, date1, date2,cell):
    
            arr = cell['Arrival']
            dep = cell['Departure'] - timedelta(1)
            code = cell['Rate code']
            
            con1 = con[(arr<=con["second date"]) & (date1>=con["first date"])]
            con2 = con[(date2<=con["second date"]) & (dep>=con["first date"])]
            
            con1['first date'].iloc[0] = arr
            con1['second date'].iloc[-1] = date1
            
            con2['first date'].iloc[0] = date2
            con2['second date'].iloc[-1] = dep
            
            result = pd.concat([con1, con2], axis=0)
            if date2 == dep:
                result = con1
            empty_df = result.empty
            days = 0 
            price = 0
            
            if not empty_df:
                days = np.array((result['second date'] - result['first date']).dt.days)
                price = np.array(result[cell['Rate code']])
            return result
            
        def invoice_optimizer(cell,invoice):
            begin_date = invoice['first date'].iloc[0]
            end_date = invoice['second date'].iloc[-1]
            
            arr = cell['Arrival']
            dep = cell['Departure'] - timedelta(1)
            code = cell['Rate code']

            if begin_date > arr:
                arr = begin_date
            if end_date < dep:
                dep = end_date
            
            date_range = invoice[(arr<=invoice["second date"]) & (dep>=invoice["first date"])]
            empty_df = date_range.empty
            
            days = 0 
            price = 0
            
            if not empty_df:
                date_range['first date'].iloc[0] = arr
                date_range['second date'].iloc[-1] = dep
                
                days = np.array((date_range['second date'] - date_range['first date']).dt.days +1)
                price = np.array(date_range[cell['Rate code']])
            
            return days, price, arr, dep, empty_df
        
        all_dis_types = [None,'reduction','extra']
        numeric_pattern = r'^[+-]?(\d*\.\d+|\d+(,\d{3})*|\d+)$'
        days = np.empty(0)
        price = np.empty(0)
        if "Spo_dict" in st.session_state:
            Spo_dict = st.session_state["Spo_dict"]
            
            con2spo_once = False
            if len(Spo_dict["name"]) > 0:
                for guest in range(len(statment['Arrival'])):
                    passing = False
                    cnt = 0
                    con2spo = False
                    contract = False
                    spo_days = np.empty(0)
                    spo_price = np.zeros(0)
                    
                    for spo_num in reversed(range(len(Spo_dict["name"]))):
                        if 'SPO' in Spo_dict:
                            SPO = Spo_dict['SPO'][spo_num].copy()
                        else:
                            SPO = pd.read_excel(st.session_state["uploaded file"],sheet_name=Spo_dict["name"][spo_num])
                        cell =  statment.iloc[guest,:]
                        threshold = 2
                        SPO = SPO.dropna(thresh=len(SPO.columns) - threshold + 1)

                        if (statment['Res_date'][guest] >= pd.Timestamp(Spo_dict['start_date'][spo_num])) and (statment['Res_date'][guest] <= pd.Timestamp(Spo_dict['end_date'][spo_num])):
                            
                            if (statment['Arrival'][guest] >= SPO['first date'][0]) and (statment['Arrival'][guest] <= SPO['second date'].iloc[-1]):
                                
                                cnt +=1
                                if cnt ==1:
                                    statment["Total price currency"][guest] = 0
                                Summing = 0
                                first_day_spo2 = pd.Timestamp(SPO["first date"][0])
                                last_day_spo2 = pd.Timestamp(SPO["second date"].iloc[-1])
                                date_arrival = pd.Timestamp(statment.loc[guest, "Arrival"])
                                if passing:
                                    date_arrival = new_arrival
                                date_departure = statment.loc[guest, "Departure"] - timedelta(1)

                                
                                res_date = statment["Res_date"][guest]
                                rate_code = statment["Rate code"][guest]
                                arrival_row = SPO[(SPO["first date"]<=date_arrival) & (SPO["second date"]>=date_arrival)]
                                
                                price_arrival_night = arrival_row[rate_code]
                                date1_arrival = pd.to_datetime(arrival_row["first date"].values[0])
                                date2_arrival = pd.to_datetime(arrival_row["second date"].values[0])
                                
                                one_spo = (date_departure<=last_day_spo2)
                                
                                
                                if (not(one_spo)):
                                    date_departure =last_day_spo2
                                departure_row = SPO[(SPO["first date"]<=date_departure) & (SPO["second date"]>=date_departure)]
                                date1_departure = pd.to_datetime(departure_row["first date"].values[0])
                                date2_departure = pd.to_datetime(departure_row["second date"].values[0])
                                
                                if date_departure <= date2_arrival:
                                    price = price_arrival_night * ((date_departure-date_arrival).days +1)
                                    price = calculate_offer(cell,Spo_dict,price,spo_num)
                                    statment["Total price currency"][guest] += price
                                    
                                    
                                else:
                                    date_range = SPO[(date_arrival<=SPO["second date"]) & (date_departure>=SPO["first date"])]
                                    diff = (date_range["second date"] - date_range["first date"]).dt.days + 1

                                    for j in range(len(date_range[rate_code])):
                                        Summing = (date_range[rate_code].iloc[j]*diff.iloc[j]) + Summing
                                    other_price = (((date_arrival-date1_arrival).days) * (arrival_row[rate_code].values[0]) + ((date2_departure-date_departure).days) * (departure_row[rate_code].values[0]))
                                    
                                    price = Summing - other_price 
                                    
                                    price = calculate_offer(cell,Spo_dict,price,spo_num)
                                    
                                    statment["Total price currency"][guest] += price 
                                if one_spo:
                                    break
                                elif not(one_spo) and (spo_num > 0):
                                    passing = True
                                    new_arrival = last_day_spo2 + timedelta(days=1)
                                else:
                                    date_arrival = last_day_spo2 + timedelta(days=1)
                                    date_departure = statment.loc[guest, "Departure"] - timedelta(1)
                                    arrival_row = con[(con["first date"]<=date_arrival) & (con["second date"]>=date_arrival)]
                
                                    departure_row = con[(con["first date"]<=date_departure) & (con["second date"]>=date_departure)]
                                    # if arrival_row.isnull().any().item():
                                    #     break
                                    
                                    price_arrival_night = arrival_row[rate_code]
                                    date1_arrival = pd.to_datetime(arrival_row["first date"].values[0])
                                    date2_arrival = pd.to_datetime(arrival_row["second date"].values[0])
                                    
                                    
                                    date1_departure = pd.to_datetime(departure_row["first date"].values[0])
                                    date2_departure = pd.to_datetime(departure_row["second date"].values[0])
                                    if date_departure <= date2_arrival:
                                        price = price_arrival_night * ((date_departure-date_arrival).days +1)
                                        # price = calculate_offer(cell,Spo_dict,price,spo_num)
                                        price = calculate_offer_con(cell,price)
                                        statment["Total price currency"][guest] += price
                                        
                                    else:
                                        date_range = con[(date_arrival<=con["second date"]) & (date_departure>=con["first date"])]
                                        diff = (date_range["second date"] - date_range["first date"]).dt.days 
                                        diff += 1
                                        for j in range(len(date_range[rate_code])):
                                            Summing = (date_range[rate_code].iloc[j]*diff.iloc[j]) + Summing
                                            
                                        other_price = (((date_arrival-date1_arrival).days) * (arrival_row[rate_code].values[0]) + ((date2_departure-date_departure).days+1) * (departure_row[rate_code].values[0]))
                                        
                                        statment.loc[guest,"other_price"] = other_price
                                        price = Summing - other_price 
                                        price = calculate_offer_con(cell,price)
                                        # price = calculate_offer(cell,Spo_dict,price,spo_num)
                                        statment.loc[guest,"Total price currency"] += price
                                    if checked1:
                                        res_date = statment["Res_date"][guest]
                                        if ((res_date >= spo_arrival_df["first date"][0]) and (res_date <= spo_arrival_df["second date"][0])) and ((date_arrival >= spo_arrival_df["first date"][0]) and (date_arrival <= spo_arrival_df["second date"][0])):
                                            night_price = spo_arrival_df[rate_code][0]
                                            
                                            nights = statment["Departure"][guest] - statment["Arrival"][guest]
                                            statment["Total price currency"][guest] += float(night_price * nights.days)
                                        elif ((res_date >= spo_arrival_df["first date"][0]) and (res_date <= spo_arrival_df["second date"][0])) and ((date_arrival >= spo_arrival_df["first date"][1]) and (date_arrival <= spo_arrival_df["second date"][1])):
                                            night_price = spo_arrival_df[rate_code][1]
                                            nights = statment["Departure"][guest] - statment["Arrival"][guest]
                                            statment["Total price currency"][guest] += float(night_price * nights.days)
                                st.write(guest)
                            elif not (statment['Arrival'][guest] >= SPO['first date'][0]):
                                con2spo = True
                                invoice_result = invoice_optimizer(cell, SPO)
                                
                                spo_days = np.append(spo_days,invoice_result[0]).reshape(-1,1)
                                spo_price = np.append(spo_price,invoice_result[1]).reshape(-1,1)
                                
                                price = calculate_offer(np.sum(np.multiply(spo_days,spo_price)))
                                statment['Total price currency'][guest] = price
                                date1 = invoice_result[2]
                                date2 = invoice_result[3]
                                
                                empty_df = invoice_result[4]
                                if isinstance(contract,bool):
                                    contract = con
                                if not empty_df:
                                    contract = contract_removal(contract,date1,date2,cell)
                                
                    if con2spo:
                        con2spo_once = True
                        spo_days = (invoice_optimizer(cell,contract)[0]).reshape(-1,1)
                        spo_price = (invoice_optimizer(cell,contract)[1]).reshape(-1,1)

                        price = calculate_offer_con(np.sum(np.multiply(spo_days,spo_price)))
                        statment['Total price currency'][guest] += price

                                
                                            
                        
                # Reductions
                if "Dis_dict" in st.session_state:
                    Dis_dict = st.session_state["Dis_dict"]
                    
                    for i in range(len(Dis_dict['type'])):
                        Dis_dict_edit = Dis_dict.copy()
                        if all_dis_types.index(Dis_dict['type'][i])==1:
                            price = -(Dis_dict['amount'][i]/100) * statment["Total price currency"][guest]
                            
                        elif all_dis_types.index(Dis_dict['type'][i])==2:
                            if 'Pax' in statment.columns:
                                pax = statment['Pax'][guest]
                            else:
                                pax = 1
                            price = Dis_dict['amount'][i] * pax
                            
                            
                        if Dis_dict['column'][i] is None:
                            statment["Total price currency"][guest] += (price)
                        elif isinstance(statment[Dis_dict['column'][i]][guest],str):
                            if statment[Dis_dict['column'][i]][guest].lower() == 'yes':
                                statment["Total price currency"][guest] += (price)
                            elif statment[Dis_dict['column'][i]][guest].lower() == 'no':
                                continue
                            elif all(statment[Dis_dict['column'][i]].str.match(numeric_pattern, na=False)):
                                statment["Total price currency"][guest] += float(statment[Dis_dict['column'][i]][guest]) * price
                            if any((statment[Dis_dict['column'][i]].str.contains(statment["Rate code"][guest], case=False, na=False))):
                                Type_of_room = statment["Rate code"][guest][0].lower()
                                
                                # adjusting room type
                                
                                type_of_room_mapping = {
                                    "s": 1,
                                    "d": 2,
                                    "t": 3,
                                    "q": 4
                                }

                                if Type_of_room in type_of_room_mapping:
                                    mapped_value = type_of_room_mapping[Type_of_room]
                                    statment["Total price currency"][guest] += mapped_value * price
                        else:
                            statment["Total price currency"][guest] += statment[Dis_dict['column'][i]][guest] * price
                        
                                
                            
                                 
                
                            
        # extra
        if "Extra" in statment.columns:
            statment["Extra"] = statment["Extra"].fillna(0)
            statment["Total price currency"] += statment['Extra']
            
        statment.drop(columns=["other_price","UnNeeded_price"],inplace=True)
        if "Amount-hotel" in statment.columns:
            statment["Difference Hotel Currency"] = statment["Total price currency"] -  statment["Amount-hotel"]   
            statment['Difference Hotel Currency'] = statment['Difference Hotel Currency'].round(1)
            if "Currency rate" in statment.columns:
                statment["Total price LE"] = statment["Total price currency"] * statment["Currency rate"]         
                statment["Difference Hotel Currency LE"] = statment["Total price LE"] -  statment["Invoice Amount L.E"] 
                statment['Difference Hotel Currency LE'] = statment['Difference Hotel Currency LE'].round(1)
        
        
        
                    
        # @st.cache_data
        # def convert_df(df):
        #     # IMPORTANT: Cache the conversion to prevent computation on every rerun
        #     return df.to_csv().encode('utf-8')
        

        # csv = convert_df(statment)

        # st.download_button(
        #     label="Download data as CSV",
        #     data=csv,
        #     file_name='answers_df.csv',
        #     mime='text/csv',
        # )
        from io import BytesIO
        from pyxlsb import open_workbook as open_xlsb
        from openpyxl.styles import PatternFill
        if 'Rate Euro.' in statment.columns:
            statment.drop(columns="Rate Euro.",inplace=True)
            
        if 'Guest Name' in statment.columns:
            statment.drop(columns="Guest Name",inplace=True)
            
        if 'Tax Date' in statment.columns:
            statment.drop(columns="Tax Date",inplace=True)
            
        if '14%  Tax.' in statment.columns:
            statment.drop(columns="14%  Tax.",inplace=True)
            
        if 'Internal Tax invoice No.' in statment.columns:
            statment.drop(columns="Internal Tax invoice No.",inplace=True)
            
        if 'Net Euro' in statment.columns:
            statment.drop(columns="Net Euro",inplace=True)
            
        if 'Currency rate' in statment.columns:
            statment.drop(columns="Currency rate",inplace=True)
            
        if 'Reservation No.' in statment.columns:
            statment.drop(columns="Reservation No.",inplace=True)
            
        if 'Description' in statment.columns:
            statment.drop(columns="Description",inplace=True)
            
        if 'System Amount' in statment.columns:
            statment.drop(columns="System Amount",inplace=True)
            
        if 'Net Amount.' in statment.columns:
            statment.drop(columns="Net Amount.",inplace=True)
            
        if 'Night' in statment.columns:
            statment.drop(columns="Night",inplace=True)
            
        if 'Adj.' in statment.columns:
            statment.drop(columns="Adj.",inplace=True)
            
        statment["Total price currency"] = statment["Total price currency"].round(2).apply(lambda x: '{:.2f}'.format(x))
        fill_pattern = PatternFill(patternType='solid',fgColor='C64747')
        import pandas as pd
        if 'Arrival' in statment.columns and pd.api.types.is_datetime64_any_dtype(statment['Arrival']):
            statment['Arrival'] = statment['Arrival'].dt.strftime('%Y/%m/%d')  # Formatting to YYYY-MM-DD

        if 'Departure' in statment.columns and pd.api.types.is_datetime64_any_dtype(statment['Departure']):
            statment['Departure'] = statment['Departure'].dt.strftime('%Y/%m/%d')  # Formatting to YYYY-MM-DD



        
        def to_excel(df):
            output = BytesIO()
            writer = pd.ExcelWriter(output, engine='xlsxwriter')
            df.to_excel(writer, index=False, sheet_name='Sheet1')
            workbook = writer.book
            worksheet = writer.sheets['Sheet1']
            format0 = workbook.add_format({
                'left': 6,
                'right': 6,  # 1 means a solid border
            })
            # Create a format with a green background color and solid border
            format1 = workbook.add_format({
                'num_format': '0.00',
                'bg_color': '99CCFF',
                'font_size': 15,   
                'bold': True,
                'font_name': 'Times New Roman',
                'left': 6,
                'right': 6,  # 1 means a solid border
                'bottom': 2,
                'align': 'center',
                'valign': 'vcenter'
            })
            format3 = workbook.add_format({
                'num_format': '0.00',
                'bg_color': 'FFFF00',
                'font_size': 15,   
                'bold': True,
                'font_name': 'Times New Roman',
                'left': 6,
                'right': 6,  # 1 means a solid border
                'bottom': 2,
                'align': 'center',
                'valign': 'vcenter'
            })
            # Set the format for cell A1
            date_format = workbook.add_format({'num_format': 'yyyy-mm-dd'})  # Change the format as per your preference
            worksheet.set_column('D:G', None, date_format)
            
            num_rows, num_cols = df.shape
            # Iterate through the cells and apply the format only to non-empty cells
            for row in range(num_rows):
                for col in range(num_cols):
                    cell_value = df.iat[row, col]
                    if pd.notna(cell_value):  # Check if the cell contains any value
                        worksheet.write(row + 1, col, cell_value, format0)  # "+1" because we skip the header row
            
            for col_num, value in enumerate(df.columns):
                worksheet.write(0, col_num, value, format1)
                
            red_format = workbook.add_format({'bg_color': '#FF0000', 'font_color': '#FFFFFF'})
            yellow_format = workbook.add_format({'bg_color': '#FFFF00', 'font_color': '#000000'})
            
            if "Currency rate" in statment.columns:
                worksheet.conditional_format(1, df.columns.get_loc("Difference Hotel Currency LE"), len(df.index), df.columns.get_loc("Difference Hotel Currency LE"),
                                            {'type': 'cell', 'criteria': '>', 'value': 0, 'format': red_format})
                worksheet.conditional_format(1, df.columns.get_loc("Difference Hotel Currency LE"), len(df.index), df.columns.get_loc("Difference Hotel Currency LE"),
                                            {'type': 'cell', 'criteria': '<', 'value': 0, 'format': yellow_format})
            if "Difference Hotel Currency" in statment.columns:
                worksheet.conditional_format(1, df.columns.get_loc("Difference Hotel Currency"), len(df.index), df.columns.get_loc("Difference Hotel Currency"),
                                            {'type': 'cell', 'criteria': '>', 'value': 0, 'format': red_format})
                worksheet.conditional_format(1, df.columns.get_loc("Difference Hotel Currency"), len(df.index), df.columns.get_loc("Difference Hotel Currency"),
                                            {'type': 'cell', 'criteria': '<', 'value': 0, 'format': yellow_format})
            if "Folio" in statment.columns:
                indices = df.index[df['Difference Hotel Currency'] > 0].tolist()

                for index in indices:
                    Folio_column_index = df.columns.get_loc('Folio')
                    cell_value = df.iat[index, Folio_column_index]
                    worksheet.write(index+1, Folio_column_index, cell_value, red_format)
                    
                indices = df.index[df['Difference Hotel Currency'] < 0].tolist()

                for index in indices:
                    Folio_column_index = df.columns.get_loc('Folio')
                    cell_value = df.iat[index, Folio_column_index]
                    worksheet.write(index+1, Folio_column_index, cell_value, yellow_format)
            if "Difference Hotel Currency" in statment.columns:
                worksheet.write(0,df.columns.get_loc("Difference Hotel Currency"), "Difference Hotel Currency", format3)
                worksheet.write(0,df.columns.get_loc("Total price currency"), "Total price currency", format3)
            if "Currency rate" in statment.columns:
                worksheet.write(0,df.columns.get_loc("Difference Hotel Currency LE"), "Difference Hotel Currency LE", format3)
                worksheet.write(0,df.columns.get_loc("Total price LE"), "Total price LE", format3)
            # HEREE
            # worksheet.set_column('A:Z', 20)
            # worksheet.set_column('L:Z', 25)
            # worksheet.set_column('O:O', 40)
            # worksheet.set_column('M:M', 34)
            worksheet.set_column('A:Z', 20)
            worksheet.set_column('A:Z', 20)
            worksheet.set_column('F:J', 27)
            worksheet.set_column('J:J', 37)
            worksheet.set_column('H:H', 34)
            
            # writer.save()
            writer.close()
            return output.getvalue(), 'Result sheet.xlsx'

        

        # Use the checkbox value
        
                
        # Assuming "statment" is your DataFrame, you can call the function like this:
        df_xlsx = to_excel(statment)
        
                
            
        if __name__ == '__main__':
            # Assuming you have your DataFrame df ready
            processed_data, file_name = to_excel(statment)

            st.download_button("Download Excel", processed_data, file_name='Result sheet.xlsx', mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        import numpy as np
        if "Difference Hotel Currency" in statment.columns:
            diffs = statment.loc[np.abs(statment["Difference Hotel Currency"])>1,:]

        
        hide_checkbox_label_style = """
            <style>
                .checkbox-container .stCheckbox>label {
                    display: none;
                }
            </style>
        """
        # Display the CSS
        # st.markdown(hide_checkbox_label_style, unsafe_allow_html=True)
        # Create the checkbox
        if "Difference Hotel Currency" in statment.columns:

            checked = st.checkbox("Show Differences",value = True)
            

            # Use the checkbox value
            if checked:
                
                st.table(diffs)
        # df_xlsx, filename = to_excel(statment)

        # # Get the current filepath
        # current_filepath = os.getcwd()
        # # Combine the current filepath with the filename to get the full file path
        # output_filepath = os.path.join(current_filepath, filename)
        # print(output_filepath)
        # # Write the Excel file to the current filepath
        # with open(r"D:\vscoded\Excl\app V.0\SPO_app-win32-x64\Result sheet.xlsx", 'wb') as f:
        #     f.write(df_xlsx)